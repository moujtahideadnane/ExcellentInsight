import os
import tempfile

import pandas as pd
import pytest

from app.pipeline.parser import (
    is_sheet_empty,
    normalize_column_name,
    parse_excel,
)


@pytest.mark.asyncio
async def test_parse_csv():
    """Test CSV parsing."""
    # Create a temporary CSV file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("name,age,city\n")
        f.write("John,30,NYC\n")
        f.write("Jane,25,LA\n")
        temp_path = f.name

    try:
        result = await parse_excel(temp_path)

        assert len(result.sheets) == 1
        assert result.sheets[0].name == "default"
        assert result.sheets[0].row_count == 2
        assert "name" in result.dataframes["default"].columns
        assert "age" in result.dataframes["default"].columns
    finally:
        os.unlink(temp_path)


@pytest.mark.asyncio
async def test_parse_multisheet_xlsx():
    """Test multi-sheet XLSX parsing."""
    # Create a temporary Excel file with multiple sheets
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name

    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            df1 = pd.DataFrame({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
            df2 = pd.DataFrame({"x": [10, 20], "y": [100, 200]})
            df1.to_excel(writer, sheet_name="Sheet1", index=False)
            df2.to_excel(writer, sheet_name="Sheet2", index=False)

        result = await parse_excel(temp_path)

        assert len(result.sheets) == 2
        sheet_names = [s.name for s in result.sheets]
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names

        # Check columns are normalized
        df1_cols = result.dataframes["Sheet1"].columns
        assert "col1" in df1_cols
        assert "col2" in df1_cols
    finally:
        os.unlink(temp_path)


@pytest.mark.asyncio
async def test_empty_sheet_is_skipped():
    """Test that empty sheets are skipped."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name

    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            df1 = pd.DataFrame({"col1": [1, 2]})
            df_empty = pd.DataFrame()  # Empty
            df1.to_excel(writer, sheet_name="Data", index=False)
            df_empty.to_excel(writer, sheet_name="Empty", index=False)

        result = await parse_excel(temp_path)

        # Should only have the Data sheet
        assert len(result.sheets) == 1
        assert result.sheets[0].name == "Data"
    finally:
        os.unlink(temp_path)


def test_normalize_column_name():
    """Test column name normalization."""
    assert normalize_column_name("First Name") == "first_name"
    assert normalize_column_name("Email-Address") == "email_address"
    assert normalize_column_name("   Spaced   ") == "spaced"
    assert normalize_column_name("") == "unnamed"
    assert normalize_column_name("Special!@#$%") == "special"


def test_is_sheet_empty():
    """Test empty sheet detection."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Empty sheet
    assert is_sheet_empty(ws) == True

    # Sheet with just headers
    ws["A1"] = "Header"
    assert is_sheet_empty(ws) == True

    # Sheet with data
    ws["A2"] = "Data"
    assert is_sheet_empty(ws) == False
