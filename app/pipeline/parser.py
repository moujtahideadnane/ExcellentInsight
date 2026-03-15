import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import polars as pl
import structlog

from app.config import get_settings

logger = structlog.get_logger(__name__)


@dataclass
class SheetMetadata:
    name: str
    row_count: int
    column_count: int
    columns: List[str]


@dataclass
class ParsedData:
    sheets: List[SheetMetadata]
    dataframes: Dict[str, pl.DataFrame]


def normalize_column_name(name: str) -> str:
    """Normalize column names: lower, snake_case, alphanumeric only."""
    if not name or str(name).strip() == "":
        return "unnamed"

    # Remove special characters, keep spaces and alphanumeric
    s = re.sub(r"[^a-zA-Z0-9\s_-]", "", str(name))
    # Replace spaces and hyphens with underscores
    s = re.sub(r"[\s-]+", "_", s.strip())
    # Lowercase
    return s.lower()


def is_year_or_numeric_only_row(row: Any) -> bool:
    """Check if a row is effectively just year/numeric labels with little else.

    We want to detect cases where the entire row (or all but one cell) are numbers
    such as a stand-alone time index row.  A row that mixes text labels with many
    numeric year columns (e.g. "Nom de la variable", "Description", 2024, 2025,
    ...) should **not** be classified as numeric-only since it is a legitimate
    header containing both identifiers and numeric column names.

    The rules here are intentionally conservative:
    * If every non-empty value can be converted to float, return True.
    * Otherwise if more than 90% of the non-empty values are numeric **and**
      there is at most one non-numeric cell, treat as numeric-only.
    * Otherwise return False.
    """
    # collect non-null, non-empty values
    values = [v for v in row if v is not None and str(v).strip() != ""]
    if not values:
        return False

    numeric_count = 0
    non_numeric_items = []

    for v in values:
        try:
            float(v)
            numeric_count += 1
        except (ValueError, TypeError):
            non_numeric_items.append(v)

    # all values numeric -> definitely numeric-only
    if numeric_count == len(values):
        return True

    # compute ratio of numeric to total
    numeric_ratio = numeric_count / len(values)
    # if almost everything is numeric but there are one or zero stray text cells,
    # still treat it as numeric-only (common when first column is blank/label)
    if numeric_ratio > 0.9 and len(non_numeric_items) <= 1:
        return True

    return False


def detect_transposed_orientation(ws: Any, header_idx: int) -> bool:
    """Detect if table is transposed (headers in first column, not first row).

    Common in financial/reporting tables:
      | Q1      | 2020  | 2021  | 2022
      | Q2      | 100   | 150   | 200
      | Revenue | 500   | 600   | 700

    Detection heuristic:
    1. First column has high text/categorical diversity (potential row labels)
    2. First row (starting from header_idx) is mostly numeric/uniform (potential column indices)
    3. Minimum data rows to consider (avoid false positives on tiny tables)

    Returns True if table appears transposed (row-oriented).
    """
    # Need at least 3+ columns and 3+ rows to consider transposition
    # Peek at first few rows to analyze
    rows_to_analyze = list(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 10, values_only=True))
    if len(rows_to_analyze) < 3:
        return False  # Too few rows to determine

    # Extract first column values (excluding header)
    first_col = []
    for row in rows_to_analyze:
        if row and row[0] is not None:
            first_col.append(str(row[0]).strip())

    # Extract first row values (the header row itself)
    header_row = list(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1, values_only=True))[0]
    if not header_row:
        return False

    # Check first column: should have good variety (text, different values)
    header_row_values = [v for v in header_row if v is not None]
    first_col_unique = set(first_col) if first_col else set()

    # Compute metrics
    first_col_diversity = len(first_col_unique) / len(first_col) if first_col else 0
    header_row_numeric = (
        sum(1 for v in header_row_values if v is not None and isinstance(v, (int, float))) / len(header_row_values)
        if header_row_values
        else 0
    )

    # Signal transposition if:
    # - First column has good categorical diversity (>60% unique)
    # - First row is mostly numeric (>60% numeric values)
    # - First column looks like labels (text, varied)
    is_first_col_varied = first_col_diversity > 0.6 and len(first_col) >= 3
    is_first_row_numeric = header_row_numeric > 0.6

    transposed = is_first_col_varied and is_first_row_numeric

    if transposed:
        logger.debug(
            "Detected transposed orientation (row headers)",
            first_col_diversity=first_col_diversity,
            header_row_numeric=header_row_numeric,
        )

    return transposed


def transpose_worksheet_data(ws: Any) -> List[List[Any]]:
    """Transpose all worksheet data: convert row-oriented (headers in column 1) to column-oriented.

    Used for financial/reporting tables where row labels are in the first column.
    Extracts all data and returns transposed 2D list.

    Returns list of rows where the original first column becomes the first row (new headers).
    """
    # Read all non-empty data
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        # Include the row even if it has some Nones (sparse data is OK)
        if any(v is not None for v in row):
            all_rows.append(list(row))

    if not all_rows:
        return []

    # Transpose: rows become columns, columns become rows
    max_cols = max(len(row) for row in all_rows) if all_rows else 0
    transposed = []

    for col_idx in range(max_cols):
        new_row = []
        for row in all_rows:
            new_row.append(row[col_idx] if col_idx < len(row) else None)
        transposed.append(new_row)

    logger.debug(
        "Transposed worksheet data",
        original_rows=len(all_rows),
        original_cols=max_cols,
        transposed_rows=len(transposed),
    )
    return transposed


def detect_header_row(ws: Any) -> int:
    """Detect the header row by finding the first row with diverse, text-like values.

    General strategy works for most Excel files:
    1. Skip empty rows
    2. Skip rows where all values are identical (section headers)
    3. Skip rows that are year-only or numeric-only (use next row instead)
    4. Pick the first row with good variety and reasonable text length

    This is format-agnostic and doesn't assume specific data patterns.
    """

    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        values = [v for v in row if v is not None]

        # Skip completely empty rows
        if not values:
            continue

        string_values = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
        if not string_values:
            continue

        # Skip rows where all cells have the same value (section/category headers)
        unique_values = set(string_values)
        if len(unique_values) <= 1:
            continue

        # Skip rows that are year-only or numeric-only (e.g., 2020, 2021, 2022, ...)
        # These are often column indices in time-series data, not actual headers
        # Instead, use the next row as the actual header
        if is_year_or_numeric_only_row(row):
            logger.debug("Skipping numeric-only header row; checking next row", row_idx=i)
            # Peek at the next row to see if it's a better header
            rows_list = list(ws.iter_rows(min_row=i + 2, max_row=i + 2, values_only=True))
            if rows_list:
                next_row = rows_list[0]
                next_values = [v for v in next_row if v is not None]
                if next_values and not is_year_or_numeric_only_row(next_row):
                    return i + 1
            continue

        # Check if row has reasonable header characteristics:
        # - Good variety (more than 1 unique value per cell count)
        # - Relatively short average value length (headers are usually concise)
        diversity_ratio = len(unique_values) / len(string_values)
        avg_length = sum(len(v) for v in string_values) / len(string_values)

        # Headers typically have good diversity and moderate length (not all huge numbers)
        if diversity_ratio > 0.3 and avg_length < 100:
            return i

    return 0  # Fallback to first row


def is_sheet_empty(ws: Any) -> bool:
    """Check if sheet has any data beyond a potential header row."""
    # If less than 2 rows (header + at least 1 data row), consider candidate for empty
    # But more accurately, check if there are any non-null values beyond row 1
    for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return False
    return True








async def save_to_parquet(job_id: str, dataframes: Dict[str, pl.DataFrame]) -> None:
    """Saves all DataFrames for a job as Parquet files for fast retrieval."""
    settings = get_settings()
    job_dir = Path(settings.STORAGE_LOCAL_PATH) / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    for name, df in dataframes.items():
        # Clean name for filesystem
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "", name)
        df.write_parquet(job_dir / f"{safe_name}.parquet")


async def load_from_parquet(job_id: str) -> Optional[ParsedData]:
    """Loads DataFrames from Parquet files if they exist."""
    settings = get_settings()
    job_dir = Path(settings.STORAGE_LOCAL_PATH) / job_id
    if not job_dir.exists():
        return None

    parquet_files = list(job_dir.glob("*.parquet"))
    if not parquet_files:
        return None

    dataframes = {}
    sheets_metadata = []

    for p in parquet_files:
        sheet_name = p.stem
        df = pl.read_parquet(p)
        dataframes[sheet_name] = df
        sheets_metadata.append(
            SheetMetadata(name=sheet_name, row_count=df.height, column_count=df.width, columns=df.columns)
        )

    return ParsedData(sheets=sheets_metadata, dataframes=dataframes)


async def parse_excel(file_path: str, job_id: Optional[str] = None) -> ParsedData:
    # 1. Check for Parquet cache first
    if job_id:
        cached = await load_from_parquet(job_id)
        if cached:
            return cached

    path = Path(file_path)
    extension = path.suffix.lower()

    dataframes = {}
    sheets_metadata = []

    if extension == ".csv":
        df = pl.read_csv(file_path)
        # Normalize columns for CSV too
        df.columns = [normalize_column_name(c) for c in df.columns]
        sheet_name = "default"
        dataframes[sheet_name] = df
        sheets_metadata.append(
            SheetMetadata(name=sheet_name, row_count=df.height, column_count=df.width, columns=df.columns)
        )
    elif extension in [".xlsx", ".xls"]:
        # 1. Use openpyxl (read_only=True) for lightweight sheet structure analysis
        # This prevents loading the entire file into RAM, saving ~90% memory
        settings = get_settings()
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames[: settings.MAX_SHEETS_PER_FILE]
        if len(wb.sheetnames) > settings.MAX_SHEETS_PER_FILE:
            # Log or raise: for now just truncate to first MAX_SHEETS_PER_FILE
            pass

        for name in sheet_names:
            ws = wb[name]

            # 2. Extract first 100 rows for smart detection (Orientation, Headers)
            # This is fast and memory-safe even on massive sheets
            preview_rows = list(ws.iter_rows(max_row=100, values_only=True))
            if not preview_rows:
                continue

            # Heuristic: Find header row
            header_idx = 0
            for i, row in enumerate(preview_rows[:20]):
                if any(v is not None for v in row) and not is_year_or_numeric_only_row(row):
                    header_idx = i
                    break

            # Check for transposed orientation
            is_transposed = detect_transposed_orientation(ws, header_idx=header_idx)

            if is_transposed:
                # Transposed sheets still require manual handling (less common)
                transposed_data = transpose_worksheet_data(ws)
                if not transposed_data: continue
                raw_headers = transposed_data[0]
                rows = transposed_data[1:]
                col_dict = {normalize_column_name(str(h)): [r[i] for r in rows] for i, h in enumerate(raw_headers)}
                df = pl.DataFrame(col_dict)
            else:
                # 3. Use HIGH PERFORMANCE Polars engine to read the actual data
                # engine="calamine" is implemented in Rust and is ultra-memory efficient
                try:
                    df = pl.read_excel(
                        file_path,
                        sheet_name=name,
                        engine="calamine",
                        read_options={"header_row": header_idx, "n_rows": settings.MAX_ROWS_PER_SHEET}
                    )
                except Exception as e:
                    logger.warning("Calamine failed, falling back to basic reader", error=str(e))
                    # Fallback to standard Polars excel reader
                    df = pl.read_excel(file_path, sheet_name=name)

            # 4. Vectorized Clean and Type Inference (Rust-speed)
            # Replace slow Python cleaning loop with Polars expressions
            if df.height > 0:
                # Normalize column names
                df.columns = [normalize_column_name(c) for c in df.columns]
                
                # Perform numeric cleanup on all columns that look numeric
                # but might be typed as Strings due to formatting ($ symbols, spaces, etc.)
                for col in df.columns:
                    if df[col].dtype == pl.String:
                        # Fast vectorized regex cleaning for numeric strings
                        # Removes spaces, currency symbols, and common group separators
                        try:
                            df = df.with_columns(
                                pl.col(col)
                                .str.replace_all(r"[^0-9\.\-eE]", "")
                                .cast(pl.Float64, strict=False)
                                .alias(f"{col}_num")
                            )
                            # If cast succeeded for enough rows, swap it
                            if df[f"{col}_num"].null_count() < (df.height * 0.8):
                                df = df.drop(col).rename({f"{col}_num": col})
                            else:
                                df = df.drop(f"{col}_num")
                        except Exception:
                            pass

                dataframes[name] = df
                sheets_metadata.append(
                    SheetMetadata(name=name, row_count=df.height, column_count=df.width, columns=df.columns)
                )

        wb.close()
    else:
        raise ValueError(f"Format de fichier non supporté : {extension}")

    # 2. Save to Parquet for next time if job_id provided
    if job_id:
        await save_to_parquet(job_id, dataframes)

    return ParsedData(sheets=sheets_metadata, dataframes=dataframes)
